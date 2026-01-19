"""Excel writing and formatting logic."""

from __future__ import annotations

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

import config


def write_to_excel(cube_data: List[Dict[str, str]], output_path: str) -> None:
    """
    Write cube data to formatted Excel workbook.

    Steps:
    1. Create workbook with "Raw" sheet containing all data
    2. Split into sheets: 45D, 60D, 45DWP, 60DWP
    3. Apply cell merging for pour locations (column G)
    4. Merge every two rows in configured columns
    """
    wb = Workbook()
    wb.active.title = config.RAW_SHEET

    create_raw_sheet(wb, cube_data)
    split_by_type(wb)

    for name in config.CONCRETE_TYPES:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for col in config.MERGE_COLS:
            merge_every_two(ws, col)
        merge_pour_locations(ws, config.POUR_LOCATION_COL)

    wb.save(output_path)


def get_concrete_type(cube_mark_prefix: str) -> str:
    """Determine concrete type from cube mark prefix."""
    mark = (cube_mark_prefix or "").upper()
    is45 = "45D" in mark
    is60 = "60D" in mark
    is_wp = "WP" in mark
    if is45 and is_wp:
        return "45DWP"
    if is60 and is_wp:
        return "60DWP"
    if is45:
        return "45D"
    if is60:
        return "60D"
    return "Unknown"


def create_raw_sheet(wb: Workbook, cube_data: List[Dict[str, str]]) -> None:
    """Create Raw sheet with all cube data."""
    ws = _create_or_clear(config.RAW_SHEET, wb)
    ws.append(config.EXCEL_HEADERS)

    for cube in cube_data:
        ws.append(
            [
                cube.get("cube_mark_prefix", ""),
                _maybe_int(cube.get("cube_number", "")),
                cube.get("cube_suffix", ""),
                cube.get("report_number", ""),
                cube.get("date_cast", ""),
                _maybe_float(cube.get("compressive_strength", "")),
                cube.get("pour_location", ""),
            ]
        )


def split_by_type(wb: Workbook) -> None:
    """Split Raw sheet into type-specific sheets."""
    src = wb[config.RAW_SHEET]
    header = [cell.value for cell in src[1]]
    targets = {t: _create_or_clear(t, wb) for t in config.CONCRETE_TYPES}
    for t_ws in targets.values():
        t_ws.append(header)

    for row in src.iter_rows(min_row=2, values_only=False):
        mark = row[0].value
        t_type = get_concrete_type(mark)
        if t_type in targets:
            targets[t_type].append([cell.value for cell in row])


def merge_pour_locations(ws: Worksheet, col_letter: str = "G") -> None:
    """Merge consecutive cells with same pour location."""
    col = ws[col_letter]
    i, last = 2, ws.max_row
    while i <= last:
        start = i
        curr = ws[f"{col_letter}{i}"].value
        while i <= last and ws[f"{col_letter}{i}"].value == curr:
            i += 1
        if i - start > 1:
            ws.merge_cells(
                start_row=start,
                start_column=col[0].column,
                end_row=i - 1,
                end_column=col[0].column,
            )
            _set_vertical_center(ws[f"{col_letter}{start}"])


def merge_every_two(ws: Worksheet, col_letter: str) -> None:
    """Merge every two rows in specified column."""
    for row in range(2, ws.max_row, 2):
        if row + 1 <= ws.max_row:
            ws.merge_cells(f"{col_letter}{row}:{col_letter}{row + 1}")
            _set_vertical_center(ws[f"{col_letter}{row}"])


def _create_or_clear(ws_name: str, wb: Workbook) -> Worksheet:
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(title=ws_name)
    return ws


def _set_vertical_center(cell) -> None:
    cell.alignment = cell.alignment.copy(vertical="center")


def _maybe_int(value):
    if value is None or value == "":
        return ""
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return value


def _maybe_float(value):
    if value is None or value == "":
        return ""
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return value
