from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from io import BytesIO
import pandas as pd

def write_to_excel(df, output):
    template_path = "concrete_cube_report_tool/templates/Concrete-strength-statistic-Superstructure.xlsx"
    wb = load_workbook(template_path)
    ws = wb.active

    start_row = 5  # Adjust as needed
    last_loc = None
    loc_start_row = start_row

    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        ws[f"B{i}"] = row.B
        ws[f"C{i}"] = row.C
        ws[f"D{i}"] = row.D
        ws[f"E{i}"] = row.E
        ws[f"F{i}"] = row.F
        ws[f"H{i}"] = float(row.H)
        ws[f"O{i}"] = row.O

        if last_loc is None:
            last_loc = row.O
            loc_start_row = i
        elif row.O != last_loc:
            if loc_start_row != i - 1:
                ws.merge_cells(f"O{loc_start_row}:O{i-1}")
                ws[f"O{loc_start_row}"].alignment = Alignment(vertical="center", wrap_text=True)
            last_loc = row.O
            loc_start_row = i

    if loc_start_row != i:
        ws.merge_cells(f"O{loc_start_row}:O{i}")
        ws[f"O{loc_start_row}"].alignment = Alignment(vertical="center", wrap_text=True)

    wb.save(output)
