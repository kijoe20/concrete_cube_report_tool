# cube_processing.py
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

RAW_SHEET = "Raw"
TARGET_SHEETS = ["45D", "60D", "45DWP", "60DWP"]

def get_strict_type(mark: str) -> str:
    """Replicates the logic of the VBA GetStrictConcreteType function."""
    mark = (mark or "").upper()
    is45 = "45D" in mark
    is60 = "60D" in mark
    is_wp = "WP" in mark
    if is45 and is_wp:
        return "45DWP"
    if is60 and is_wp:
        return "60DWP"
    if is45:
        return "45D"
    if is60:
        return "60D"
    return "Unknown"

def create_or_clear(ws_name: str, wb: Workbook) -> Worksheet:
    """Create a new worksheet or clear an existing one while preserving structure."""
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
        ws.delete_rows(1, ws.max_row)        # Clear contents (preserve sheet)
    else:
        ws = wb.create_sheet(title=ws_name)
    return ws

def split_by_type(wb: Workbook) -> None:
    """Split raw data into separate worksheets by concrete type."""
    src = wb[RAW_SHEET]
    # Read header once for efficiency
    header = [cell.value for cell in src[1]]
    # Create four target worksheets and copy header
    targets = {t: create_or_clear(t, wb) for t in TARGET_SHEETS}
    for t_ws in targets.values():
        t_ws.append(header)
    # Process each row and distribute to appropriate worksheet
    for r in src.iter_rows(min_row=2, values_only=False):
        mark = r[0].value                       # Column A
        t_type = get_strict_type(mark)
        if t_type in targets:
            targets[t_type].append([cell.value for cell in r])

def merge_pour_locations(ws: Worksheet, col_letter: str = "G") -> None:
    """Merge consecutive cells with the same value in pour location column (G) vertically."""
    col = ws[col_letter]
    i, last = 2, ws.max_row
    while i <= last:
        start = i
        curr = ws[f"{col_letter}{i}"].value
        # Find consecutive cells with same value
        while i <= last and ws[f"{col_letter}{i}"].value == curr:
            i += 1
        if i - start > 1:
            ws.merge_cells(start_row=start, start_column=col[0].column,
                           end_row=i-1,  end_column=col[0].column)
            cell = ws[f"{col_letter}{start}"]
            cell.alignment = cell.alignment.copy(vertical="center")

def merge_every_two(ws: Worksheet, col_letter: str) -> None:
    """Merge every two rows vertically in specified column (A/B)."""
    for row in range(2, ws.max_row, 2):
        if row+1 <= ws.max_row:
            ws.merge_cells(f"{col_letter}{row}:{col_letter}{row+1}")
            cell = ws[f"{col_letter}{row}"]
            cell.alignment = cell.alignment.copy(vertical="center")

def run_all() -> None:
    """Main execution function that orchestrates the entire processing workflow."""
    INPUT_FILE = "templates/cube_data_06.xlsx"
    OUTPUT_FILE = "cube_data_processed.xlsx"
    
    try:
        wb = load_workbook(INPUT_FILE)
    except FileNotFoundError:
        print(f"Error: Input file '{INPUT_FILE}' not found.")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    try:
        split_by_type(wb)

        for name in TARGET_SHEETS:
            ws = wb[name]
            merge_pour_locations(ws, "G")
            merge_every_two(ws, "A")
            merge_every_two(ws, "B")

        wb.save(OUTPUT_FILE)
        print(f"Processing complete. Output saved to '{OUTPUT_FILE}'.")
        
    except Exception as e:
        print(f"Error during processing: {e}")
        return

if __name__ == "__main__":
    run_all()